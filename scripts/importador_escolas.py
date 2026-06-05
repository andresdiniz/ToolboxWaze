import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import pandas as pd
import time
import os
from datetime import datetime
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

# ─── Paleta de cores ────────────────────────────────────────────────────────
BG       = "#1e1e2e"
SURFACE  = "#2a2a3e"
SURFACE2 = "#313145"
ACCENT   = "#4f98a3"
SUCCESS  = "#6daa45"
WARNING  = "#e8af34"
ERROR    = "#dd6974"
TEXT     = "#cdccca"
MUTED    = "#797876"
WHITE    = "#f0efed"
FONT_BODY  = ("Segoe UI", 10)
FONT_BOLD  = ("Segoe UI", 10, "bold")
FONT_TITLE = ("Segoe UI", 13, "bold")
FONT_MONO  = ("Consolas", 9)

# Colunas candidatas para ID da escola (ordem de prioridade)
# Cobre tanto o CSV do MEC (Código INEP) quanto outros formatos
ID_COLS = [
    "Código INEP",   # CSV padrão MEC (Análise Detalhada)
    "CO_ENTIDADE",   # Microdados MEC
    "CO_ESCOLA",
    "ID_ESCOLA",
    "CO_INEP",
    "codigo_inep",
    "CODIGO_INEP",
    "ID",
]

# Colunas candidatas para nome da escola
NOME_COLS = [
    "Escola",        # CSV padrão MEC
    "NO_ENTIDADE",   # Microdados MEC
    "NOME_ESCOLA",
    "nome_escola",
    "NM_ESCOLA",
]

# Colunas candidatas para município
MUNICIPIO_COLS = [
    "Município",     # CSV padrão MEC
    "NO_MUNICIPIO",
    "MUNICIPIO",
]

# Colunas candidatas para UF
UF_COLS = [
    "UF",            # CSV padrão MEC
    "SG_UF",
    "sg_uf",
]

# Colunas candidatas para latitude e longitude
LAT_COLS = ["Latitude",  "NU_LATITUDE",  "lat", "latitude"]
LON_COLS = ["Longitude", "NU_LONGITUDE", "lon", "longitude"]


def _primeira_col(row_dict: dict, candidatas: list, fallback: str = "") -> str:
    """Retorna o valor da primeira coluna encontrada no dicionário."""
    for col in candidatas:
        val = str(row_dict.get(col, "")).strip()
        if val and val.lower() not in ("nan", "none", ""):
            return val
    return fallback


def extrair_id(row_dict: dict) -> str:
    return _primeira_col(row_dict, ID_COLS, fallback="—")


def extrair_nome(row_dict: dict, idx: int) -> str:
    return _primeira_col(row_dict, NOME_COLS, fallback=f"Escola #{idx}")


def extrair_municipio(row_dict: dict) -> str:
    return _primeira_col(row_dict, MUNICIPIO_COLS)


def extrair_uf(row_dict: dict) -> str:
    return _primeira_col(row_dict, UF_COLS)


def extrair_coords(row_dict: dict):
    """Retorna (lat, lon) como float se válidos, ou (None, None)."""
    try:
        lat = float(_primeira_col(row_dict, LAT_COLS, "").replace(",", "."))
        lon = float(_primeira_col(row_dict, LON_COLS, "").replace(",", "."))
        if lat and lon:
            return lat, lon
    except (ValueError, TypeError):
        pass
    return None, None


# ─── Geocodificador com estratégia em camadas ────────────────────────────────
class GeocoderCamadas:
    def __init__(self):
        self.geo = Nominatim(user_agent="escola_geocoder_br/1.0", timeout=8)
        self._delay = 1.1

    def _tentar(self, query: str):
        try:
            time.sleep(self._delay)
            loc = self.geo.geocode(query, country_codes="br", language="pt")
            return (loc.latitude, loc.longitude) if loc else None
        except (GeocoderTimedOut, GeocoderServiceError):
            return None

    def geocodificar(self, row_dict: dict):
        municipio = extrair_municipio(row_dict)
        uf        = extrair_uf(row_dict)
        nome      = _primeira_col(row_dict, NOME_COLS, "")

        # Campos de endereço (microdados MEC)
        logr   = str(row_dict.get("DS_ENDERECO", "")).strip()
        numero = str(row_dict.get("NU_ENDERECO", "")).strip()
        bairro = str(row_dict.get("NO_BAIRRO",   "")).strip()

        # Camada 1 — endereço completo
        partes = [p for p in [logr, numero, bairro, municipio, uf, "Brasil"]
                  if p and p.lower() not in ("nan", "none", "")]
        if partes:
            res = self._tentar(", ".join(partes))
            if res:
                return {"lat": res[0], "lon": res[1], "estrategia": "endereco_completo", "status": "ok"}

        # Camada 2 — nome da escola + município + UF
        if nome and municipio:
            res = self._tentar(f"{nome}, {municipio}, {uf}, Brasil")
            if res:
                return {"lat": res[0], "lon": res[1], "estrategia": "nome_escola", "status": "ok"}

        # Camada 3 — centroide do município
        if municipio:
            res = self._tentar(f"{municipio}, {uf}, Brasil")
            if res:
                return {"lat": res[0], "lon": res[1], "estrategia": "centroide_municipio", "status": "parcial"}

        return {"lat": None, "lon": None, "estrategia": "—", "status": "sem_coords"}


# ─── Janela principal ────────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Importador de Escolas — Geocodificação em Camadas")
        self.geometry("1160x740")
        self.minsize(820, 560)
        self.configure(bg=BG)
        self.resizable(True, True)

        self._df_result = pd.DataFrame()
        self._running   = False
        self._stop_flag = threading.Event()
        self._geocoder  = GeocoderCamadas()

        self._build_ui()
        self._log("Sistema iniciado. Aguardando arquivo CSV.", "info")

    # ── UI ──────────────────────────────────────────────────────────────────
    def _build_ui(self):
        tk.Label(self, text="📍 Importador de Escolas com Geocodificação",
                 bg=BG, fg=ACCENT, font=FONT_TITLE).pack(pady=(14, 4))
        tk.Label(self, text="Estratégia em camadas: endereço completo → nome+município → centroide",
                 bg=BG, fg=MUTED, font=FONT_BODY).pack(pady=(0, 10))

        ff = tk.Frame(self, bg=SURFACE, padx=10, pady=8)
        ff.pack(fill="x", padx=16, pady=(0, 8))
        tk.Label(ff, text="Arquivo CSV:", bg=SURFACE, fg=TEXT, font=FONT_BOLD).pack(side="left")
        self._path_var = tk.StringVar(value="Nenhum arquivo selecionado")
        tk.Label(ff, textvariable=self._path_var, bg=SURFACE, fg=MUTED,
                 font=FONT_BODY, anchor="w").pack(side="left", fill="x", expand=True, padx=8)
        self._btn_browse = self._btn(ff, "📂 Abrir CSV", self._browse, ACCENT)
        self._btn_browse.pack(side="right")

        pf = tk.Frame(self, bg=BG)
        pf.pack(fill="x", padx=16, pady=(0, 6))
        self._prog_label = tk.Label(pf, text="Aguardando...", bg=BG, fg=MUTED, font=FONT_BODY)
        self._prog_label.pack(anchor="w")
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Geo.Horizontal.TProgressbar",
                        troughcolor=SURFACE2, background=ACCENT,
                        bordercolor=SURFACE2, lightcolor=ACCENT, darkcolor=ACCENT)
        self._progress = ttk.Progressbar(pf, style="Geo.Horizontal.TProgressbar",
                                          orient="horizontal", mode="determinate")
        self._progress.pack(fill="x", pady=(2, 0))

        bf = tk.Frame(self, bg=BG)
        bf.pack(fill="x", padx=16, pady=(0, 8))
        self._btn_start  = self._btn(bf, "▶  Iniciar",  self._start,  SUCCESS)
        self._btn_stop   = self._btn(bf, "⏹  Parar",    self._stop,   ERROR,   state="disabled")
        self._btn_export = self._btn(bf, "💾 Exportar", self._export, WARNING, state="disabled")
        self._btn_clear  = self._btn(bf, "🗑 Limpar",   self._clear,  MUTED)
        for b in [self._btn_start, self._btn_stop, self._btn_export, self._btn_clear]:
            b.pack(side="left", padx=4)

        paned = tk.PanedWindow(self, orient="vertical", bg=SURFACE2, sashwidth=5, sashrelief="flat")
        paned.pack(fill="both", expand=True, padx=16, pady=(0, 14))

        lf = tk.Frame(paned, bg=SURFACE)
        tk.Label(lf, text=" Log de Execução", bg=SURFACE2, fg=ACCENT,
                 font=FONT_BOLD, anchor="w").pack(fill="x")
        self._log_text = tk.Text(lf, bg=SURFACE, fg=TEXT, font=FONT_MONO,
                                  state="disabled", wrap="word", relief="flat",
                                  insertbackground=TEXT, selectbackground=ACCENT)
        sb_l = tk.Scrollbar(lf, command=self._log_text.yview, bg=SURFACE2)
        self._log_text.configure(yscrollcommand=sb_l.set)
        sb_l.pack(side="right", fill="y")
        self._log_text.pack(fill="both", expand=True, padx=4, pady=4)
        for tag, fg in [("ok", SUCCESS), ("warn", WARNING), ("error", ERROR),
                        ("info", TEXT), ("accent", ACCENT), ("muted", MUTED),
                        ("id", "#c9a0dc")]:
            self._log_text.tag_config(tag, foreground=fg)
        paned.add(lf, minsize=140)

        tf = tk.Frame(paned, bg=SURFACE)
        tk.Label(tf, text=" Resultados", bg=SURFACE2, fg=ACCENT,
                 font=FONT_BOLD, anchor="w").pack(fill="x")
        cols = ("ID INEP", "Escola", "Município", "UF", "Estratégia", "Lat", "Lon", "Status")
        self._tree = ttk.Treeview(tf, columns=cols, show="headings", height=8)
        style.configure("Treeview", background=SURFACE, foreground=TEXT,
                         fieldbackground=SURFACE, rowheight=22, font=FONT_BODY)
        style.configure("Treeview.Heading", background=SURFACE2,
                         foreground=ACCENT, font=FONT_BOLD)
        style.map("Treeview", background=[("selected", ACCENT)],
                   foreground=[("selected", WHITE)])
        widths  = {"ID INEP": 100, "Escola": 250, "Município": 120, "UF": 40,
                   "Estratégia": 155, "Lat": 100, "Lon": 100, "Status": 90}
        anchors = {"ID INEP": "center", "Escola": "w", "Município": "w",
                   "UF": "center", "Estratégia": "w",
                   "Lat": "center", "Lon": "center", "Status": "center"}
        for c in cols:
            self._tree.heading(c, text=c)
            self._tree.column(c, width=widths[c], anchor=anchors[c])
        self._tree.tag_configure("ok",      foreground=SUCCESS)
        self._tree.tag_configure("parcial", foreground=WARNING)
        self._tree.tag_configure("sem",     foreground=ERROR)
        sb_tv = ttk.Scrollbar(tf, orient="vertical",   command=self._tree.yview)
        sb_tx = ttk.Scrollbar(tf, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=sb_tv.set, xscrollcommand=sb_tx.set)
        sb_tv.pack(side="right", fill="y")
        sb_tx.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True, padx=4, pady=4)
        paned.add(tf, minsize=120)

    def _btn(self, parent, text, cmd, color, state="normal"):
        return tk.Button(parent, text=text, command=cmd, bg=color, fg=WHITE,
                         font=FONT_BOLD, relief="flat", padx=14, pady=6,
                         activebackground=ACCENT, activeforeground=WHITE,
                         cursor="hand2", state=state)

    # ── Log ─────────────────────────────────────────────────────────────────
    def _log(self, msg: str, kind: str = "info"):
        ts   = datetime.now().strftime("%H:%M:%S")
        icon = {"ok":"✔","warn":"⚠","error":"✖","info":"ℹ","accent":"►","muted":"·","id":"#"}.get(kind,"·")
        self._log_text.configure(state="normal")
        self._log_text.insert("end", f"[{ts}] {icon}  {msg}\n", kind)
        self._log_text.see("end")
        self._log_text.configure(state="disabled")

    def _log_escola(self, idx: int, total: int, escola_id: str, nome: str, kind: str, detalhe: str):
        """Log padronizado com ID INEP destacado em roxo claro."""
        ts   = datetime.now().strftime("%H:%M:%S")
        icon = {"ok":"✔","warn":"⚠","error":"✖","accent":"►"}.get(kind,"·")
        id_tag = f"ID {escola_id}"
        linha  = f"[{ts}] {icon}  [{idx:>5}/{total}] {id_tag:<12}  {nome[:45]:<45}  {detalhe}\n"
        self._log_text.configure(state="normal")
        insert_pos = self._log_text.index("end")
        self._log_text.insert("end", linha, kind)
        # Destaca o ID INEP em roxo claro
        full = self._log_text.get(f"{insert_pos} linestart", f"{insert_pos} lineend")
        start = full.find(id_tag)
        if start != -1:
            base = self._log_text.index(f"{insert_pos} linestart")
            self._log_text.tag_add("id",
                                   f"{base} + {start} chars",
                                   f"{base} + {start + len(id_tag)} chars")
        self._log_text.see("end")
        self._log_text.configure(state="disabled")

    # ── Arquivo ─────────────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if path:
            self._path_var.set(path)
            self._log(f"Arquivo selecionado: {os.path.basename(path)}", "accent")

    # ── Importação ──────────────────────────────────────────────────────────
    def _start(self):
        path = self._path_var.get()
        if not os.path.isfile(path):
            messagebox.showerror("Erro", "Selecione um arquivo CSV válido.")
            return
        self._stop_flag.clear()
        self._running = True
        self._btn_start.config(state="disabled")
        self._btn_stop.config(state="normal")
        self._btn_export.config(state="disabled")
        for item in self._tree.get_children():
            self._tree.delete(item)
        threading.Thread(target=self._run, args=(path,), daemon=True).start()

    def _stop(self):
        self._stop_flag.set()
        self._log("Interrupção solicitada — aguardando escola atual...", "warn")

    def _run(self, path: str):
        # Tenta UTF-8 com BOM (padrão MEC) e cai em latin-1 se falhar
        for enc in ("utf-8-sig", "latin-1"):
            try:
                df = pd.read_csv(path, sep=",", encoding=enc, low_memory=False, dtype=str)
                if len(df.columns) < 2:
                    df = pd.read_csv(path, sep=";", encoding=enc, low_memory=False, dtype=str)
                df.columns = [c.strip() for c in df.columns]
                break
            except Exception:
                continue
        else:
            self._log("Erro ao ler CSV — verifique o separador e a codificação.", "error")
            self._finish()
            return

        total = len(df)
        cols_detectadas = list(df.columns[:8])
        self._log(f"CSV carregado — {total} registros | colunas detectadas: {cols_detectadas}", "info")

        # Informa qual coluna de ID foi encontrada
        col_id_encontrada = next((c for c in ID_COLS if c in df.columns), None)
        if col_id_encontrada:
            self._log(f"Coluna de ID INEP detectada: '{col_id_encontrada}'", "ok")
        else:
            self._log("⚠ Nenhuma coluna de ID INEP encontrada — usando índice de linha como fallback", "warn")

        self._progress["maximum"] = total
        resultados = []
        stats = {"ok": 0, "parcial": 0, "sem_coords": 0, "ignorado": 0}

        for i, (_, row) in enumerate(df.iterrows(), 1):
            if self._stop_flag.is_set():
                self._log("Processamento interrompido pelo usuário.", "warn")
                break

            row_dict  = row.to_dict()
            escola_id = extrair_id(row_dict) if col_id_encontrada else f"linha-{i}"
            nome      = extrair_nome(row_dict, i)
            municipio = extrair_municipio(row_dict)
            uf        = extrair_uf(row_dict)
            situacao  = str(row_dict.get("TP_SITUACAO_FUNCIONAMENTO", "1")).strip()

            self.after(0, self._update_prog, i, total, int(i / total * 100), escola_id, nome)

            # Já tem coordenadas salvas?
            lat, lon = extrair_coords(row_dict)
            if lat and lon:
                res = {"lat": lat, "lon": lon, "estrategia": "coords_existentes", "status": "ok"}
                stats["ok"] += 1
                self._log_escola(i, total, escola_id, nome, "ok",
                                 f"coords existentes → {lat:.5f}, {lon:.5f}")
                self._add_row(escola_id, nome, municipio, uf, res)
                resultados.append({**row_dict, **res})
                if i % 50 == 0:
                    self._log_resumo(i, total, stats)
                continue

            # Paralisada / extinta — ignora geocodificação
            if situacao not in ("1", ""):
                res = {"lat": None, "lon": None, "estrategia": "—", "status": "ignorado"}
                stats["ignorado"] += 1
                self._log_escola(i, total, escola_id, nome, "muted",
                                 f"situação={situacao} → ignorado")
                resultados.append({**row_dict, **res})
                continue

            # Geocodificação em camadas
            self._log_escola(i, total, escola_id, nome, "accent", "geocodificando...")
            res = self._geocoder.geocodificar(row_dict)

            if res["status"] == "ok":
                stats["ok"] += 1
                self._log_escola(i, total, escola_id, nome, "ok",
                                 f"✔ {res['estrategia']} → {res['lat']:.5f}, {res['lon']:.5f}")
            elif res["status"] == "parcial":
                stats["parcial"] += 1
                self._log_escola(i, total, escola_id, nome, "warn",
                                 f"⚠ centroide município → {res['lat']:.5f}, {res['lon']:.5f}")
            else:
                stats["sem_coords"] += 1
                self._log_escola(i, total, escola_id, nome, "error",
                                 "✖ sem_coords — revisar manualmente")

            self._add_row(escola_id, nome, municipio, uf, res)
            resultados.append({**row_dict, **res})

            if i % 50 == 0:
                self._log_resumo(i, total, stats)

        self._df_result = pd.DataFrame(resultados)
        self._log("─" * 70, "muted")
        self._log(
            f"Concluído! {total} registros | "
            f"✔ ok={stats['ok']}  ⚠ parcial={stats['parcial']}  "
            f"✖ sem_coords={stats['sem_coords']}  · ignorados={stats['ignorado']}",
            "ok"
        )
        self._finish()

    def _log_resumo(self, i, total, stats):
        self._log(
            f"── {i}/{total} processados | "
            f"ok={stats['ok']}  parcial={stats['parcial']}  "
            f"sem={stats['sem_coords']}  ignorado={stats['ignorado']}",
            "muted"
        )

    def _update_prog(self, i, total, pct, escola_id, nome):
        self._progress["value"] = i
        self._prog_label.config(
            text=f"{i}/{total} ({pct}%)  ID {escola_id}  —  {nome[:50]}",
            fg=ACCENT
        )

    def _add_row(self, escola_id, nome, municipio, uf, res):
        tag = {"ok": "ok", "parcial": "parcial",
               "sem_coords": "sem", "ignorado": "sem"}.get(res["status"], "sem")
        lat = f"{res['lat']:.6f}" if res["lat"] else "—"
        lon = f"{res['lon']:.6f}" if res["lon"] else "—"
        self.after(0, self._tree.insert, "", "end",
                   values=(escola_id, nome[:50], municipio, uf,
                            res["estrategia"], lat, lon, res["status"]),
                   tags=(tag,))

    def _finish(self):
        self._running = False
        self.after(0, lambda: self._btn_start.config(state="normal"))
        self.after(0, lambda: self._btn_stop.config(state="disabled"))
        if not self._df_result.empty:
            self.after(0, lambda: self._btn_export.config(state="normal"))
        self.after(0, lambda: self._prog_label.config(text="Concluído.", fg=SUCCESS))

    # ── Exportar ────────────────────────────────────────────────────────────
    def _export(self):
        if self._df_result.empty:
            messagebox.showinfo("Aviso", "Nenhum dado para exportar.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            initialfile=f"escolas_{datetime.now().strftime('%Y%m%d_%H%M')}"
        )
        if not path:
            return
        try:
            if path.endswith(".csv"):
                self._df_result.to_csv(path, index=False, sep=";", encoding="utf-8-sig")
            else:
                with pd.ExcelWriter(path, engine="openpyxl") as writer:
                    self._df_result.to_excel(writer, index=False, sheet_name="Escolas")
                    wb = writer.book
                    ws = wb["Escolas"]
                    from openpyxl.styles import PatternFill, Font, Alignment
                    hf    = PatternFill("solid", fgColor="1F3864")
                    hfont = Font(bold=True, color="FFFFFF", name="Calibri")
                    for cell in ws[1]:
                        cell.fill = hf
                        cell.font = hfont
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions
                    for col in ws.columns:
                        w = max(len(str(c.value or "")) for c in col) + 3
                        ws.column_dimensions[col[0].column_letter].width = min(w, 50)
            self._log(f"Exportado com sucesso: {os.path.basename(path)}", "ok")
        except Exception as e:
            self._log(f"Erro ao exportar: {e}", "error")

    # ── Limpar ──────────────────────────────────────────────────────────────
    def _clear(self):
        self._log_text.configure(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.configure(state="disabled")
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._df_result = pd.DataFrame()
        self._progress["value"] = 0
        self._prog_label.config(text="Aguardando...", fg=MUTED)
        self._btn_export.config(state="disabled")
        self._log("Log e tabela limpos.", "muted")


if __name__ == "__main__":
    app = App()
    app.mainloop()
